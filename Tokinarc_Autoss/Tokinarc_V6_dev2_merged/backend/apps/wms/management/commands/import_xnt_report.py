"""
Tokinarc V6 — apps/wms/management/commands/import_xnt_report.py

Nạp "BÁO CÁO TỔNG HỢP NHẬP XUẤT TỒN" (kế toán, export MISA dạng .xlsx) vào
WMS thật, thay cho dữ liệu Nhập/Xuất kho test đang có:
  - Cột "Tồn đầu kỳ"    -> 1 phiếu NHẬP KHO lớn (receive_stock, ghi StockMovement thật).
  - Cột "Xuất trong kỳ" -> 1 phiếu XUẤT KHO lớn (generate_pick_list + confirm_pick_and_ship).
  Tháng này báo cáo KHÔNG có cột "Nhập trong kỳ" (không phát sinh mua hàng),
  nên "tồn đầu kỳ" đóng vai trò tồn khởi tạo cho hệ thống mới.

Mã hàng trong file kế toán (VD "T74805", "01916A-10") KHÔNG cùng hệ với mã
catalog nội bộ hiện có (VD "001001") — 2 hệ mã không liên quan, không thể tự
động khớp theo mã. Khớp theo TÊN đã chuẩn hoá (lower + strip khoảng trắng)
với Part.display_name_vi; không khớp -> tự tạo Part mới (mã XNT-NNN).

Kho/Ô/NCC không có trong báo cáo (chỉ là số liệu tổng cả tháng) -> chọn NGẪU
NHIÊN trong dữ liệu có sẵn (kho/ô đang có, NCC đang có) theo yêu cầu.

Mặc định CHỈ XEM TRƯỚC (dry-run), không ghi gì:
    python manage.py import_xnt_report "đường dẫn file.xlsx"
Ghi thật (xoá sạch Inbound/Outbound/Inventory cũ rồi nạp lại):
    python manage.py import_xnt_report "đường dẫn file.xlsx" --yes --user quanly1
"""
from __future__ import annotations

import random
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.accounts.models import User
from apps.catalog.models import Part
from apps.purchasing.models import Supplier
from apps.wms import services
from apps.wms.models import (
    Bin, InboundLine, InboundOrder, InventoryItem, Lot, OutboundLine,
    OutboundOrder, SerialNumber, StockMovement, Warehouse,
)


def _norm(s: str) -> str:
    return ' '.join((s or '').strip().lower().split())


class Command(BaseCommand):
    help = ('Nạp báo cáo Nhập-Xuất-Tồn (kế toán, .xlsx) vào WMS: Tồn đầu kỳ -> phiếu '
            'Nhập kho, Xuất trong kỳ -> phiếu Xuất kho. Thay cho dữ liệu test hiện có.')

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path')
        parser.add_argument('--yes', action='store_true',
                             help='Xác nhận ghi thật. Không có cờ này chỉ in ra sẽ làm gì, không ghi.')
        parser.add_argument('--user', default='',
                             help='Username gán làm người thao tác (mặc định: admin).')

    def handle(self, xlsx_path, yes, user, **opts):
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        rows = []
        for r in range(11, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            # Chỉ nhận dòng hàng THẬT (có STT số) — bỏ dòng "Tổng cộng"/chữ ký
            # cuối báo cáo (cột B của mấy dòng đó có chữ nên "if not name" không lọc được).
            if not isinstance(stt, (int, float)) or not name:
                continue
            rows.append({
                'name': str(name).strip(),
                'dvt': ws.cell(r, 3).value,
                'ton_dau_luong': ws.cell(r, 4).value or 0,
                'ton_dau_tien': ws.cell(r, 5).value or 0,
                'xuat_luong': ws.cell(r, 6).value or 0,
                'xuat_tien': ws.cell(r, 7).value or 0,
                'ton_cuoi_luong': ws.cell(r, 8).value or 0,
            })
        self.stdout.write(f"Đọc được {len(rows)} dòng hàng từ {xlsx_path}.")

        # ── Khớp Part theo tên (đã chuẩn hoá); không khớp -> sẽ tạo mới ──
        existing_by_name = {_norm(n): p for n, p in Part.objects.values_list('display_name_vi', 'tokin_part_no')}
        last_xnt = (Part.objects.filter(tokin_part_no__startswith='XNT-')
                    .order_by('-tokin_part_no').values_list('tokin_part_no', flat=True).first())
        next_seq = int(last_xnt.split('-')[1]) + 1 if last_xnt else 1

        matched, to_create = [], []
        for row in rows:
            code = existing_by_name.get(_norm(row['name']))
            if code:
                row['part_no'] = code
                matched.append(row)
            else:
                row['part_no'] = f'XNT-{next_seq:03d}'
                next_seq += 1
                to_create.append(row)

        total_ton_dau = sum(r['ton_dau_luong'] for r in rows)
        total_xuat = sum(r['xuat_luong'] for r in rows)
        self.stdout.write(f"Khớp catalog hiện có: {len(matched)} dòng. Sẽ tạo Part mới: {len(to_create)} dòng.")
        self.stdout.write(f"Tổng Tồn đầu kỳ: {total_ton_dau:,.0f}. Tổng Xuất trong kỳ: {total_xuat:,.0f}.")

        if not yes:
            self.stdout.write(self.style.WARNING(
                "\nCHƯA ghi gì cả (thiếu --yes). Xem lại số liệu trên rồi chạy lại kèm --yes để ghi thật."))
            self.stdout.write("Vài Part sẽ tạo mới (mẫu):")
            for r in to_create[:10]:
                self.stdout.write(f"  {r['part_no']}: {r['name']}")
            return

        actor = User.objects.filter(username=user).first() if user else None
        actor = actor or User.objects.filter(username='admin').first() or User.objects.filter(is_admin=True).first()
        if actor is None:
            raise CommandError("Không tìm thấy user để gán người thao tác (thử --user <username>).")

        # Chọn ngẫu nhiên trong các kho active THỰC SỰ có ô (bin) — nhiều server
        # có kho mới tạo chưa xếp ô (VD "HANOI"), .first() theo mã dễ trúng kho rỗng.
        candidate_whs = [w for w in Warehouse.objects.filter(is_active=True)
                         if Bin.objects.filter(zone__warehouse=w).exists()]
        if not candidate_whs:
            raise CommandError("Không có kho active nào có ô (bin) để nạp dữ liệu.")
        wh = random.choice(candidate_whs)
        bins = list(Bin.objects.filter(zone__warehouse=wh))
        supplier_names = list(Supplier.objects.filter(is_active=True).values_list('name', flat=True)) or ['NCC (chưa rõ)']

        with transaction.atomic():
            # ── Xoá sạch Inbound/Outbound/Inventory cũ (giống reset_wms_ledger) ──
            StockMovement.objects.all().delete()
            InventoryItem.objects.all().delete()
            Lot.objects.all().delete()
            SerialNumber.objects.all().delete()
            OutboundOrder.objects.all().delete()
            InboundOrder.objects.all().delete()

            # ── Tạo Part mới cho các dòng không khớp catalog hiện có ──
            for r in to_create:
                unit_cost = int(r['ton_dau_tien'] / r['ton_dau_luong']) if r['ton_dau_luong'] else 0
                Part.objects.create(
                    tokin_part_no=r['part_no'], category='Nhập từ báo cáo XNT T7',
                    display_name_vi=r['name'], price_unit=r['dvt'] or 'cái',
                    cost_vnd=unit_cost or None,
                    source='import_xnt_report',
                )

            part_by_no = {p.tokin_part_no: p for p in Part.objects.filter(
                tokin_part_no__in=[r['part_no'] for r in rows])}
            bin_by_part: dict[str, Bin] = {}

            inb = InboundOrder.objects.create(
                code=self._next_code(InboundOrder, 'IN'), warehouse=wh, status='draft',
                flow_type='internal', supplier=random.choice(supplier_names),
                notes='Nạp tồn đầu kỳ từ báo cáo Nhập-Xuất-Tồn tháng 7/2026 (kế toán).',
                created_by=actor, updated_by=actor)
            out = OutboundOrder.objects.create(
                code=self._next_code(OutboundOrder, 'OUT'), warehouse=wh, rule='FIFO', purpose='sale',
                status='draft', notes='Xuất trong kỳ từ báo cáo Nhập-Xuất-Tồn tháng 7/2026 (kế toán).',
                created_by=actor, updated_by=actor)

            idx_in = idx_out = 0
            for r in rows:
                part = part_by_no[r['part_no']]
                b = bin_by_part.setdefault(r['part_no'], random.choice(bins))
                qty_in = int(r['ton_dau_luong'] or 0)
                unit_cost = int(r['ton_dau_tien'] / qty_in) if qty_in else 0
                if qty_in > 0:
                    InboundLine.objects.create(
                        inbound=inb, part=part, qty_expected=qty_in, target_bin=b,
                        qty_received=qty_in, qty_putaway=qty_in, unit_cost=unit_cost, order_idx=idx_in)
                    idx_in += 1
                    services.receive_stock(bin_obj=b, part=part, qty=qty_in, user=actor, ref_id=inb.code)
                qty_out = int(r['xuat_luong'] or 0)
                if qty_out > 0:
                    OutboundLine.objects.create(
                        outbound=out, part=part, qty_ordered=qty_out, order_idx=idx_out)
                    idx_out += 1

            inb.status = 'putaway'
            inb.received_by = actor
            inb.save(update_fields=['status', 'received_by'])

            if idx_out:
                services.generate_pick_list(out)
                services.confirm_pick_and_ship(out, user=actor)
            else:
                out.delete()

        self.stdout.write(self.style.SUCCESS(
            f"Đã nạp xong vào kho {wh.code}: {len(to_create)} Part mới, phiếu nhập {inb.code} ({idx_in} dòng), "
            + (f"phiếu xuất {out.code} ({idx_out} dòng)." if idx_out else "không có dòng xuất.")))
        self.stdout.write("Chạy `python manage.py reconcile_stock` để xác nhận Tồn kho khớp Ledger.")

    def _next_code(self, model, prefix):
        from datetime import date
        pre = f"{prefix}-{date.today().year}-"
        last = model.objects.filter(code__startswith=pre).order_by('-code').first()
        n = int(last.code.rsplit('-', 1)[-1]) + 1 if last else 1
        return f"{pre}{n:03d}"
