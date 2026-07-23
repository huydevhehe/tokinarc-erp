"""
Tokinarc — apps/purchasing/imports.py

Import Nhà cung cấp (Supplier) hàng loạt từ Excel/CSV.
  - POST /api/v1/purchasing/suppliers/import/          (file=..., dry_run=1 xem trước)
  - GET  /api/v1/purchasing/suppliers/import-template/ (tải file Excel mẫu)

Quy tắc:
  - Chỉ Quản lý kho trở lên (PO_WRITE_ROLES) được import.
  - Có `code` → cập nhật NCC trùng mã (không tạo trùng). Thiếu `code` nhưng
    trùng `tax_code` (MST) đã có → cập nhật NCC đó. Còn lại → tạo mới, tự sinh
    mã NCC-XXXX.
  - dry_run: chỉ kiểm tra + trả thống kê/lỗi, KHÔNG ghi DB.
"""
from __future__ import annotations

import csv
import io
import re

from django.db import transaction
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.roles import role_of

from .models import Supplier
from .views import PO_WRITE_ROLES

COLUMNS = ['code', 'name', 'tax_code', 'phone', 'email', 'address', 'notes']
_CODE_NUM_RE = re.compile(r'NCC-(\d+)', re.IGNORECASE)


def _norm(s: str) -> str:
    return (s or '').strip().lower().replace(' ', '_').replace('-', '_')


def _parse_file(f) -> list[dict]:
    name = (getattr(f, 'name', '') or '').lower()
    if name.endswith('.csv'):
        text = f.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        return [{_norm(k): (v or '').strip() for k, v in row.items()} for row in reader]
    from openpyxl import load_workbook
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [_norm(str(h)) if h is not None else '' for h in next(rows)]
    except StopIteration:
        return []
    out = []
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        out.append({headers[i]: (str(c).strip() if c is not None else '')
                    for i, c in enumerate(r) if i < len(headers)})
    return out


def _next_code_seq() -> int:
    mx = 0
    for code in Supplier.objects.values_list('code', flat=True):
        m = _CODE_NUM_RE.search(code or '')
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def validate_and_plan(rows: list[dict]):
    """Trả (plan, errors). plan: list dict {mode: create|update, code, defaults}.
    Không ghi DB. mode chỉ để hiển thị preview."""
    plan, errors = [], []
    seen_codes = set()
    existing_codes = set(Supplier.objects.values_list('code', flat=True))
    tax_to_code = {t: c for c, t in Supplier.objects.exclude(tax_code='')
                   .values_list('code', 'tax_code')}

    for i, row in enumerate(rows, start=2):
        name = (row.get('name') or '').strip()
        code = (row.get('code') or '').strip()
        tax = (row.get('tax_code') or '').strip()
        # Dòng trống hoàn toàn (mọi cột rỗng) → bỏ qua; còn có bất kỳ dữ liệu nào
        # mà thiếu tên → báo lỗi (tránh nuốt im dòng nhập sót tên).
        if not any((row.get(c) or '').strip() for c in COLUMNS):
            continue
        if not name:
            errors.append({'row': i, 'message': 'Thiếu tên nhà cung cấp (name).'})
            continue

        # Xác định NCC đích: theo code → theo MST → tạo mới.
        if code:
            if code in seen_codes:
                errors.append({'row': i, 'message': f'{code}: trùng mã ngay trong file.'})
                continue
            seen_codes.add(code)
            mode = 'update' if code in existing_codes else 'create'
        elif tax and tax in tax_to_code:
            code = tax_to_code[tax]
            mode = 'update'
        else:
            mode = 'create'   # code sẽ tự sinh khi ghi thật

        defaults = {'name': name}
        for field in ('tax_code', 'phone', 'email', 'address', 'notes'):
            if row.get(field):
                defaults[field] = row[field].strip()
        plan.append({'mode': mode, 'code': code, 'defaults': defaults})
    return plan, errors


class SupplierImportView(APIView):
    """Import NCC từ Excel/CSV. ?dry_run=1 để xem trước (không ghi)."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if role_of(request.user) not in PO_WRITE_ROLES:
            return Response({'detail': 'Chỉ Quản lý kho/quản lý/CEO được import nhà cung cấp.'},
                            status=403)
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'Thiếu file.'}, status=400)
        try:
            rows = _parse_file(f)
        except Exception as e:   # noqa: BLE001
            return Response({'detail': f'Không đọc được file: {e}'}, status=400)

        plan, errors = validate_and_plan(rows)
        will_create = sum(1 for p in plan if p['mode'] == 'create')
        will_update = sum(1 for p in plan if p['mode'] == 'update')
        # dry_run có thể tới từ query (?dry_run=1, cách FE gọi) hoặc form field.
        raw_dry = request.query_params.get('dry_run') or request.data.get('dry_run') or ''
        dry = str(raw_dry).lower() in ('1', 'true', 'yes')

        if dry:
            return Response({
                'dry_run': True, 'total_rows': len(rows),
                'will_create': will_create, 'will_update': will_update, 'errors': errors,
                'preview': [{'code': p['code'] or '(tự sinh)', 'name': p['defaults']['name'],
                             'update': p['mode'] == 'update'} for p in plan[:10]],
            })

        seq = _next_code_seq()
        with transaction.atomic():
            for p in plan:
                code = p['code']
                if not code:
                    code = f'NCC-{seq:04d}'
                    seq += 1
                Supplier.objects.update_or_create(
                    code=code, defaults={**p['defaults'],
                                         'created_by': request.user, 'updated_by': request.user})
        return Response({'dry_run': False, 'created': will_create, 'updated': will_update,
                         'errors': errors})


class SupplierImportTemplateView(APIView):
    """Tải file Excel mẫu (đúng cột) để nạp NCC hàng loạt."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = 'NhaCungCap'
        ws.append(COLUMNS)
        ws.append(['', 'Công ty TNHH ABC', '0312345678', '0901234567',
                   'kinhdoanh@abc.vn', '12 Lê Lợi, Q1, TP.HCM', 'NCC phụ tùng chính'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="mau_import_nha_cung_cap.xlsx"'
        return resp
