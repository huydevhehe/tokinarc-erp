"""
Tokinarc V6 — apps/catalog/imports.py

Import danh mục phụ tùng (Part) hàng loạt từ Excel/CSV.
  - POST /api/v1/catalog/parts/import/          (file=..., dry_run=1 để xem trước)
  - GET  /api/v1/catalog/parts/import-template/ (tải file Excel mẫu)

Quy tắc:
  - Chỉ manager/CEO/admin được import (nghiệp vụ nhạy cảm, ghi hàng loạt).
  - Trùng tokin_part_no (đã tồn tại) → CẬP NHẬT các field có giá trị trong dòng
    (không tạo trùng, không xoá field cũ nếu dòng để trống).
  - dry_run: chỉ kiểm tra + trả thống kê/lỗi, KHÔNG ghi DB.
"""
from __future__ import annotations

import csv
import io

from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.roles import WMS_OP_ROLES, role_of

from .models import Part

COLUMNS = ['tokin_part_no', 'category', 'ecosystem', 'display_name_vi', 'display_name_en',
           'price_vnd', 'tax_pct', 'price_unit', 'notes']


def _norm(s: str) -> str:
    return (s or '').strip().lower().replace(' ', '_').replace('-', '_')


def _parse_file(f) -> list[dict]:
    """Đọc Excel (.xlsx) hoặc CSV → list dict {header: value}."""
    name = (getattr(f, 'name', '') or '').lower()
    if name.endswith('.csv'):
        text = f.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        return [{_norm(k): (v or '').strip() for k, v in row.items()} for row in reader]
    from openpyxl import load_workbook
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [_norm(str(h)) if h is not None else '' for h in next(rows)]
    except StopIteration:
        return []
    out = []
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        out.append({headers[i]: (str(c).strip() if c is not None else '')
                    for i, c in enumerate(r) if i < len(headers)})
    return out


def _parse_decimal(raw: str):
    raw = (raw or '').strip()
    if not raw:
        return None
    try:
        return float(raw.replace(',', ''))
    except ValueError:
        return '__invalid__'


def validate_and_build(rows: list[dict]):
    """Trả (valid_items, errors). Không ghi DB. valid_items có is_update để biết
    tạo mới hay cập nhật (chỉ để hiển thị preview, DB tự update_or_create)."""
    valid, errors = [], []
    seen = set()
    existing = set(Part.objects.values_list('tokin_part_no', flat=True))

    for i, row in enumerate(rows, start=2):   # dòng 1 = header
        code = (row.get('tokin_part_no') or '').strip()
        name = (row.get('display_name_vi') or '').strip()
        category = (row.get('category') or '').strip()
        if not code and not name:
            continue   # dòng trống
        if not code:
            errors.append({'row': i, 'message': 'Thiếu mã phụ tùng (tokin_part_no).'}); continue
        if not category:
            errors.append({'row': i, 'message': f'{code}: thiếu nhóm hàng (category).'}); continue
        if not name:
            errors.append({'row': i, 'message': f'{code}: thiếu tên (display_name_vi).'}); continue
        if code in seen:
            errors.append({'row': i, 'message': f'{code}: trùng mã ngay trong file import.'}); continue
        seen.add(code)

        price_vnd = _parse_decimal(row.get('price_vnd'))
        tax_pct = _parse_decimal(row.get('tax_pct'))
        if price_vnd == '__invalid__':
            errors.append({'row': i, 'message': f'{code}: price_vnd không phải số.'}); continue
        if tax_pct == '__invalid__':
            errors.append({'row': i, 'message': f'{code}: tax_pct không phải số.'}); continue

        defaults = {'category': category, 'display_name_vi': name}
        if row.get('ecosystem'):
            defaults['ecosystem'] = row['ecosystem'].strip()
        if row.get('display_name_en'):
            defaults['display_name_en'] = row['display_name_en'].strip()
        if price_vnd is not None:
            defaults['price_vnd'] = price_vnd
        if tax_pct is not None:
            defaults['tax_pct'] = tax_pct
        if row.get('price_unit'):
            defaults['price_unit'] = row['price_unit'].strip()
        if row.get('notes'):
            defaults['notes'] = row['notes'].strip()

        valid.append({'code': code, 'is_update': code in existing, 'defaults': defaults})
    return valid, errors


class PartImportView(APIView):
    """Import danh mục Part từ Excel/CSV. ?dry_run=1 để xem trước (không ghi)."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        # #3 biên bản (2026-07-22): mở thêm cho Kho (NV kho + QL kho) — danh mục
        # phụ tùng là dữ liệu Kho trực tiếp quản lý, không chỉ Quản lý/CEO.
        if role_of(request.user) not in WMS_OP_ROLES:
            return Response({'detail': 'Chỉ Kho/quản lý/CEO/admin được import dữ liệu.'}, status=403)
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'Thiếu file.'}, status=400)
        try:
            rows = _parse_file(f)
        except Exception as e:   # noqa: BLE001 — báo lỗi đọc file thân thiện
            return Response({'detail': f'Không đọc được file: {e}'}, status=400)

        valid, errors = validate_and_build(rows)
        dry = str(request.query_params.get('dry_run', '')).lower() in ('1', 'true', 'yes')
        will_create = sum(1 for v in valid if not v['is_update'])
        will_update = sum(1 for v in valid if v['is_update'])

        if dry:
            return Response({
                'dry_run': True, 'total_rows': len(rows),
                'will_create': will_create, 'will_update': will_update,
                'errors': errors,
                'preview': [{'code': v['code'], 'name': v['defaults'].get('display_name_vi'),
                            'update': v['is_update']} for v in valid[:10]],
            })

        with transaction.atomic():
            for v in valid:
                Part.objects.update_or_create(tokin_part_no=v['code'], defaults=v['defaults'])
        return Response({'dry_run': False, 'created': will_create, 'updated': will_update,
                         'skipped_existing': 0, 'errors': errors})


class PartImportTemplateView(APIView):
    """Tải file Excel mẫu (đúng cột) để nạp danh mục phụ tùng hàng loạt."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = 'PhuTung'
        ws.append(COLUMNS)
        ws.append(['P-99001', 'Tip hàn CO2', 'Panasonic', 'Đầu tip hàn CO2 1.2mm',
                   'CO2 Welding Tip', '15000', '10', 'cái', 'Nhập từ dữ liệu cũ'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="mau_import_phu_tung.xlsx"'
        return resp
